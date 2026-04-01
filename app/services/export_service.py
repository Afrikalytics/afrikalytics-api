"""
Export service — generates PDF, Excel, and CSV files from study/insight/report data.

Supports three formats:
  - PDF via reportlab (lightweight, no system deps)
  - Excel via openpyxl (already used for import)
  - CSV via stdlib csv module
"""

import csv
import io
import logging
from datetime import datetime, timezone
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

# Content types per format
CONTENT_TYPES = {
    "pdf": "application/pdf",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv; charset=utf-8",
}


def export_study_pdf(title: str, data: list[dict[str, Any]], columns: list[str]) -> bytes:
    """Generate a PDF report for a study dataset."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    elements: list = []

    # Header
    elements.append(Paragraph(f"Afrikalytics — {title}", styles["Title"]))
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(
        f"Export du {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 1 * cm))

    # Data table
    if data and columns:
        header = columns[:10]  # limit columns for PDF readability
        table_data = [header]
        for row in data[:500]:  # limit rows for PDF
            table_data.append([str(row.get(col, ""))[:50] for col in header])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("Aucune donnée disponible.", styles["Normal"]))

    # Footer
    elements.append(Spacer(1, 1 * cm))
    elements.append(Paragraph(
        "© 2026 Afrikalytics by Marketym — Intelligence d'Affaires pour l'Afrique",
        styles["Normal"],
    ))

    doc.build(elements)
    return buf.getvalue()


def export_study_xlsx(title: str, data: list[dict[str, Any]], columns: list[str]) -> bytes:
    """Generate an Excel workbook for a study dataset."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    # Write header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    # Write data
    for row_idx, row in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    # Auto-size columns (approximate)
    for col_idx, col_name in enumerate(columns, 1):
        max_len = max(len(str(col_name)), 10)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    # Metadata sheet
    meta = wb.create_sheet("Info")
    meta.cell(row=1, column=1, value="Titre").font = Font(bold=True)
    meta.cell(row=1, column=2, value=title)
    meta.cell(row=2, column=1, value="Export").font = Font(bold=True)
    meta.cell(row=2, column=2, value=datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC"))
    meta.cell(row=3, column=1, value="Lignes").font = Font(bold=True)
    meta.cell(row=3, column=2, value=len(data))
    meta.cell(row=4, column=1, value="Source").font = Font(bold=True)
    meta.cell(row=4, column=2, value="Afrikalytics — afrikalytics.com")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_study_csv(data: list[dict[str, Any]], columns: list[str]) -> bytes:
    """Generate a CSV file for a study dataset."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def get_content_type(fmt: str) -> str:
    """Return the MIME content type for the given export format."""
    return CONTENT_TYPES.get(fmt, "application/octet-stream")


def get_file_extension(fmt: str) -> str:
    """Return the file extension for the given export format."""
    return fmt
